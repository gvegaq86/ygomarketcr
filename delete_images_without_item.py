from helper.tyt_utils_old import TYTUtils
from helper.utils import Utils
from openpyxl import load_workbook
from helper.wcapi_utils import WCAPIUtils
wcapi = WCAPIUtils()

utils = Utils()
tyt = TYTUtils(exchange_rate=585)


loc = ("/Users/giovanni.vega/Desktop/inventory_images.xlsx")
wb = load_workbook(loc)
ws = wb['Sheet1']
ws['A1'] = 'A1'

ws_inventory_images = ws


def find_url(url_to_look):
    rows = int(ws_inventory_images.max_row)

    for i in range(0, rows - 1):
        url = ws_inventory_images[i + 2][0].value

        if url_to_look == url:
            return True
    return False

# Give the location of the file
loc = ("/Users/giovanni.vega/Desktop/media_files.xlsx")

wb = load_workbook(loc)
ws = wb['Sheet1']
ws['A1'] = 'A1'

rows = int(ws.max_row)

for i in range(0, rows-1):
    url = ws[i + 2][1].value

    if find_url(url):
        wcapi.delete_image(id=ws[i + 2][0].value)


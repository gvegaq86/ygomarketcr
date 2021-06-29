import os
import requests

from helper.tyt_utils_old import TYTUtils
from helper.utils import Utils

utils = Utils()
tyt = TYTUtils(exchange_rate=585)


def get_images(set_code, edition, condition):
    keywords = ''
    language = 'English'
    website = 'TYT'
    hide_oos = False

    tyt_card_list = tyt.get_prices(set_code=set_code, edition=edition, condition=condition, language=language,
                                   hide_oos=hide_oos)
    card_list = tyt_card_list

    if card_list:
        card_image = card_list[0]['image']
        r = requests.get(card_image, allow_redirects=True)
        open(f'card_images/{set_code}-{edition}-{condition}.jpg', 'wb').write(r.content)
        return True
    else:
        return False

# Give the location of the file
loc = ("C:\ygomarketcr\inventory2.xlsx")
from openpyxl import load_workbook

wb = load_workbook(loc)
ws = wb['inventory2']
ws['A1'] = 'A1'

rows = int(ws.max_row)

for i in range(0, rows-1):
    n = ws[i + 2][3].value
    imagenes = ws[i + 2][29].value
    codigo = ws[i + 2][40].value
    rareza = ws[i + 2][52].value
    edition = ws[i + 2][44].value
    #ws[i + 2][48].value = codigo.split("-")[0]
    condition = ws[i + 2][56].value
    #ws[i + 2][3].value = f"{codigo} {n} - {edition} - {condition} - {rareza}"
    ws[i + 2][29].value = f"https://ygomarketcr.com/wp-content/uploads/2021/07/{codigo}-{edition.replace(' ', '-')}-{condition}.jpg"
    wb.save(loc)

    item_found = False

    if codigo is not None:
        for file_name in os.listdir('C:\ygomarketcr\card_images'):
            s = file_name
            a = file_name.split("-")
            code2 = f'{a[0]}-{a[1]}'
            expansion = code2.split("-")[0]

            if a[2] == "1st":
                edition2 = "1st Edition"
            elif a[2] == "Limited":
                edition2 = "Limited Edition"
            else:
                edition2 = "Unlimited"

            print(a)
            condition2 = a[-1].replace(".jpg", "")

            if codigo == code2 and edition == edition2 and condition == condition2:
                item_found = True
                break

        if not item_found:
            if edition == None:
                edition = ""

            if condition == None:
                condition = ""

            a = get_images(set_code=codigo, edition=edition, condition=condition)
            if a:
               print(f"Imagen Descargada para el codigo {codigo}")
            else:
                print(f"Imagen NO fue descargada para el codigo {codigo}")
